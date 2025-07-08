
import streamlit as st
import pandas as pd
import datetime
import uuid
import os
from openpyxl import load_workbook
from st_aggrid import AgGrid, GridOptionsBuilder
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
import pydeck as pdk

st.set_page_config(layout="wide")
st.title("📍 Full EBar Dashboard with AI Scheduling")

EBAR_LIST = [f"EBAR{str(i).zfill(3)}" for i in range(1, 51)]
VENUES = {
    "Depot": "B1 1AA",
    "LFC": "L4 0TH",
    "EFC": "L4 4EL",
    "Crystal Palace": "SE25 6PU",
    "Edgbaston": "B5 7QU",
    "Thirst": "YO1 7HH",
    "Bradford": "BD1 1NN",
    "Ally Pally": "N22 7AY",
    "Bath": "BA2 6LP",
    "Doncaster": "DN4 5JW",
    "DCCC": "S2 1TA",
    "Trent Bridge": "NG2 6AG",
}
EVENTS_FILE = "events.csv"

geolocator = Nominatim(user_agent="ebar_dashboard")
def get_coords(postcode):
    try:
        loc = geolocator.geocode(postcode)
        return (loc.latitude, loc.longitude) if loc else None
    except:
        return None

# Load Deployment Plan
wb = load_workbook("2025 Deployment planning sheet.xlsx", data_only=True)
ws = wb["Deployment Plan 2025"]
date_row = list(ws.iter_rows(min_row=3, max_row=3))[0]
date_map = {cell.column: cell.value for cell in date_row if isinstance(cell.value, datetime.datetime)}

records = []
for row in ws.iter_rows(min_row=4):
    unit_number = str(row[1].value).zfill(3) if row[1].value else None
    ebar_id = f"EBAR{unit_number}" if unit_number else None
    for cell in row[2:]:
        col_idx = cell.column
        if col_idx not in date_map:
            continue
        date_val = date_map[col_idx].date()
        location = cell.value if cell.value else ""
        if ebar_id:
            records.append({
                "assigned_ebar": ebar_id,
                "date": date_val,
                "location": location
            })

deployment_df = pd.DataFrame(records)
latest_locations = deployment_df.sort_values("date").groupby("assigned_ebar").last().reset_index()
latest_locations = latest_locations[["assigned_ebar", "location"]].rename(columns={"location": "current_location"})
availability_map = {
    ebar: set(group["date"].tolist()) for ebar, group in deployment_df.groupby("assigned_ebar")
}

if os.path.exists(EVENTS_FILE):
    events_df = pd.read_csv(EVENTS_FILE)
    for col in ["delivery_date", "event_date", "collection_date"]:
        events_df[col] = pd.to_datetime(events_df[col], errors='coerce')
else:
    events_df = pd.DataFrame(columns=[
        "event_id", "event_name", "venue", "location",
        "delivery_date", "event_date", "collection_date", "assigned_ebar"
    ])

# --- Calendar View ---
st.header("📅 Calendar Overview")
today = datetime.date.today()
start_of_month = today.replace(day=1)
end_of_year = datetime.date(today.year, 12, 31)
calendar_days = pd.date_range(start=start_of_month, end=end_of_year)

for month, group in events_df.groupby(events_df["event_date"].dt.to_period("M")):
    st.subheader(f"{month.strftime('%B %Y')}")
    days = pd.date_range(start=month.start_time, end=month.end_time)
    for day in days:
        count = len(events_df[events_df["event_date"] == pd.Timestamp(day)])
        if count:
            with st.expander(f"{day.strftime('%A, %d %B')} — {count} EBars"):
                rows = events_df[events_df["event_date"] == pd.Timestamp(day)]
                for _, row in rows.iterrows():
                    st.markdown(f"**{row['event_name']}** at *{row['venue']}* — `{row['assigned_ebar']}`")

# --- AI Event Scheduler ---
st.header("➕ AI-Assisted Event Scheduling")
with st.form("ai_schedule"):
    event_name = st.text_input("Event Name")
    venue = st.selectbox("Venue", list(VENUES.keys()))
    location = VENUES[venue]
    delivery_date = st.date_input("Delivery Date", min_value=today)
    event_date = st.date_input("Event Date", min_value=delivery_date)
    use_collection = st.checkbox("Specify Collection Date")
    collection_date = st.date_input("Collection Date", min_value=event_date) if use_collection else None
    num_required = st.number_input("Number of EBars", min_value=1, max_value=50, step=1)

    auto_assigned = []
    if venue and delivery_date and event_date:
        block_start = delivery_date
        block_end = collection_date or event_date
        date_range = pd.date_range(block_start, block_end).date
        venue_coords = get_coords(location)

        valid_units = []
        for ebar in EBAR_LIST:
            booked = availability_map.get(ebar, set())
            current_loc = latest_locations.loc[latest_locations["assigned_ebar"] == ebar, "current_location"].values[0]
            conflict_event = False
            for i in range(1, 3):
                day = event_date + datetime.timedelta(days=i)
                if (ebar in events_df["assigned_ebar"].values and
                    ((events_df["event_date"] == pd.Timestamp(day)) &
                     (events_df["venue"] == current_loc)).any()):
                    conflict_event = True
                    break
            if any(day in booked for day in date_range) or conflict_event:
                continue
            current_coords = get_coords(VENUES.get(current_loc, current_loc))
            distance_km = geodesic(current_coords, venue_coords).km if current_coords and venue_coords else 9999
            valid_units.append((ebar, distance_km))

        ranked = sorted(valid_units, key=lambda x: x[1])
        auto_assigned = [ebar for ebar, _ in ranked[:num_required]]

    st.markdown(f"**Suggested Units:** `{', '.join(auto_assigned)}`")
    assigned_ebars = st.multiselect("Assigned EBars", EBAR_LIST, default=auto_assigned)
    submit = st.form_submit_button("Add Event")
    if submit and assigned_ebars:
        for ebar in assigned_ebars:
            row = {
                "event_id": str(uuid.uuid4()),
                "event_name": event_name,
                "venue": venue,
                "location": location,
                "delivery_date": delivery_date,
                "event_date": event_date,
                "collection_date": collection_date if use_collection else pd.NaT,
                "assigned_ebar": ebar
            }
            events_df = pd.concat([events_df, pd.DataFrame([row])], ignore_index=True)
        events_df.to_csv(EVENTS_FILE, index=False)
        st.success("Event scheduled.")

# --- EBar Schedule Table ---
st.header("📋 EBar Schedule Table")
start_date = datetime.date(2025, 7, 1)
end_date = datetime.date(2025, 12, 31)
full_days = pd.date_range(start_date, end_date)
schedule = events_df.pivot_table(index="assigned_ebar", columns="event_date", values="venue", aggfunc=lambda x: ', '.join(x), fill_value="")
schedule = schedule.reindex(index=EBAR_LIST, columns=full_days)
schedule.columns = schedule.columns.strftime("%d-%m")
schedule.reset_index(inplace=True)

gb = GridOptionsBuilder.from_dataframe(schedule)
gb.configure_default_column(resizable=True, wrapText=True, autoHeight=True)
grid_options = gb.build()
st.markdown("<style>.ag-header { background-color: black; color: white; }</style>", unsafe_allow_html=True)
AgGrid(schedule, gridOptions=grid_options, fit_columns_on_grid_load=True, height=500)

# --- Venue Map ---
st.header("🗺️ Map of EBar Deployments")
locations = []
for _, row in events_df.iterrows():
    coords = get_coords(row["location"])
    if coords:
        locations.append({
            "lat": coords[0],
            "lon": coords[1],
            "event": row["event_name"],
            "venue": row["venue"],
            "ebar": row["assigned_ebar"]
        })
if locations:
    df = pd.DataFrame(locations)
    st.pydeck_chart(pdk.Deck(
        map_style="mapbox://styles/mapbox/light-v9",
        initial_view_state=pdk.ViewState(latitude=df["lat"].mean(), longitude=df["lon"].mean(), zoom=5),
        layers=[pdk.Layer("ScatterplotLayer", data=df, get_position="[lon, lat]", get_color="[200, 30, 0, 160]", get_radius=5000)],
        tooltip={"text": "{event} at {venue} ({ebar})"}
    ))
else:
    st.info("No locations available to plot.")
